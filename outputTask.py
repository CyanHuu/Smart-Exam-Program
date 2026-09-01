import re
import shutil
from copy import copy as copy_style

import xlrd
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
from xlutils.copy import copy


def _save_workbook(workbook, path):
    """把 Windows 文件占用错误转换成用户能看懂的提示。"""
    try:
        workbook.save(path)
    except PermissionError as exc:
        raise PermissionError(
            f"无法保存文件：{path}\n"
            "该文件可能正在被 Excel/WPS 打开，请先关闭它，或换一个新的文件名后重试。"
        ) from exc


def _find_column(headers, keywords):
    # 关键词顺序代表业务优先级：先匹配实际教室编号，再匹配考试系统内部的考场号。
    for keyword in keywords:
        for index, header in enumerate(headers):
            text = str(header).strip().lower()
            if keyword in text:
                return index
    return None


def _room_key(value):
    return str(value or "").strip().casefold()


def _pad_row(row, width):
    """Excel 中允许数据行比表头短，统一补齐后再按列读取。"""
    values = list(row)
    return values + [""] * max(0, width - len(values))


def _assignment_text(teachers):
    # 用真实换行分隔人员，配合 wrap on，Excel/WPS 中不会把名单挤在一行。
    return "\n".join(
        f"{teacher[1]}（第一监考）" if index == 0 else str(teacher[1])
        for index, teacher in enumerate(teachers)
    )


def _staff_cell_style():
    return xlwt.easyxf(
        "align: horiz center, vert center, wrap on;"
        "borders: left thin, right thin, top thin, bottom thin"
    )


def _prepare_staff_column(sheet, staff_col):
    """给人员列留出可换行的宽度，保留原模板更宽的设置。"""
    column = sheet.col(staff_col)
    column.width = max(column.width or 0, 6000)


def _set_staff_row_height(sheet, row_index, value, col_index=None, merged_cells=None):
    lines = max(1, str(value or "").count("\n") + 1)
    row_start, row_end = row_index, row_index + 1
    if col_index is not None:
        for rlo, rhi, clo, chi in (merged_cells or []):
            if rlo <= row_index < rhi and clo <= col_index < chi:
                row_start, row_end = rlo, rhi
                break
    span = row_end - row_start
    height = max(420, (300 * lines + span - 1) // span)
    for row in range(row_start, row_end):
        sheet.row(row).height = height


def _apply_xlsx_staff_border(sheet, target):
    edge = Side(style="thin", color="000000")
    for merged in sheet.merged_cells.ranges:
        if not (merged.min_row <= target.row <= merged.max_row and merged.min_col <= target.column <= merged.max_col):
            continue
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                cell = sheet.cell(row, column)
                cell.border = Border(
                    left=edge if column == merged.min_col else Side(),
                    right=edge if column == merged.max_col else Side(),
                    top=edge if row == merged.min_row else Side(),
                    bottom=edge if row == merged.max_row else Side(),
                )
        return
    target.border = Border(left=edge, right=edge, top=edge, bottom=edge)


def _set_xlsx_staff_cell(sheet, target, value, add_border=True):
    target.value = value
    target.alignment = target.alignment.copy(wrap_text=True, vertical="center")
    if add_border:
        _apply_xlsx_staff_border(sheet, target)
    lines = max(1, str(value or "").count("\n") + 1)
    sheet.row_dimensions[target.row].height = max(22, min(120, 18 * lines))


def _copy_xlsx_cell(source, target):
    if source.has_style:
        # 不能直接复用源工作簿的 StyleArray；其中的字体/边框索引属于源工作簿，
        # 在目标工作簿处理合并单元格时会触发 IndexError。逐项复制可让 openpyxl
        # 在目标工作簿中重新注册这些样式。
        target.font = copy_style(source.font)
        target.fill = copy_style(source.fill)
        target.border = copy_style(source.border)
        target.alignment = copy_style(source.alignment)
        target.protection = copy_style(source.protection)
    target.number_format = source.number_format
    if source.hyperlink:
        target._hyperlink = copy_style(source.hyperlink)
    if source.comment:
        target.comment = copy_style(source.comment)


def _copy_xlsx_dimensions(source, target):
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden
        target.column_dimensions[key].bestFit = dimension.bestFit
    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key].height = dimension.height
        target.row_dimensions[key].hidden = dimension.hidden
        target.row_dimensions[key].outlineLevel = dimension.outlineLevel
    target.sheet_format.defaultColWidth = source.sheet_format.defaultColWidth
    target.sheet_format.defaultRowHeight = source.sheet_format.defaultRowHeight
    target.freeze_panes = source.freeze_panes


def write_assignments_to_excel(assignments, input_output_path, header_row=2, output_path=None):
    """把安排写入模板。

    旧调用仍可传入同一个路径；GUI 可通过 output_path 指定新的保存路径。
    """
    target_path = output_path or input_output_path
    if output_path and output_path != input_output_path:
        try:
            shutil.copyfile(input_output_path, output_path)
        except PermissionError as exc:
            raise PermissionError(
                f"无法覆盖文件：{output_path}\n"
                "该文件可能正在被 Excel/WPS 打开，请先关闭它，或换一个新的文件名后重试。"
            ) from exc

    rb = xlrd.open_workbook(target_path, formatting_info=True)
    wb = copy(rb)
    for sheet_index in range(rb.nsheets):
        sheet_rb = rb.sheet_by_index(sheet_index)
        sheet_wb = wb.get_sheet(sheet_index)
        headers = sheet_rb.row_values(header_row)
        room_col = _find_column(headers, ["教室编号", "考场号", "room"])
        staff_col = _find_column(headers, ["监考人员", "监考教师", "staff", "proctor"])
        if room_col is None or staff_col is None:
            print(f"⚠️ 跳过 sheet {sheet_index}: 未找到考场或监考人员列")
            continue

        _prepare_staff_column(sheet_wb, staff_col)
        staff_style = _staff_cell_style()
        for row_index in range(header_row + 1, sheet_rb.nrows):
            room = _merged_or_cell_value(sheet_rb, row_index, room_col)
            if room and room in assignments:
                staff_text = _assignment_text(assignments[room])
                sheet_wb.write(row_index, staff_col, staff_text, staff_style)
                _set_staff_row_height(
                    sheet_wb, row_index, staff_text, staff_col, sheet_rb.merged_cells
                )
    _save_workbook(wb, target_path)
    print(f"[OK] 监考安排已写入: {target_path}")


def write_session_assignments_to_excel(schedule_results, input_output_path, header_row=2, output_path=None, add_staff_border=True):
    """按场次分别写入模板的多个工作表，并保留原始格式。"""
    target_path = output_path or input_output_path
    if str(input_output_path).lower().endswith(".xlsx"):
        return write_session_assignments_to_xlsx(schedule_results, input_output_path, header_row, output_path, add_staff_border)
    if output_path and output_path != input_output_path:
        try:
            shutil.copyfile(input_output_path, output_path)
        except PermissionError as exc:
            raise PermissionError(
                f"无法覆盖文件：{output_path}\n"
                "该文件可能正在被 Excel/WPS 打开，请先关闭它，或换一个新的文件名后重试。"
            ) from exc
    rb = xlrd.open_workbook(target_path, formatting_info=True)
    wb = copy(rb)
    for sheet_index in range(rb.nsheets):
        sheet_rb = rb.sheet_by_index(sheet_index)
        sheet_wb = wb.get_sheet(sheet_index)
        session_id = str(sheet_rb.name)
        result = schedule_results.get(session_id)
        if not result:
            continue
        headers = sheet_rb.row_values(header_row)
        room_col = _find_column(headers, ["教室编号", "考场号", "room"])
        staff_col = _find_column(headers, ["监考人员", "监考教师", "staff", "proctor"])
        if room_col is None or staff_col is None:
            raise ValueError(f"工作表 {session_id} 未找到教室编号或监考人员列")
        assignments = result["assignments"]
        _prepare_staff_column(sheet_wb, staff_col)
        staff_style = _staff_cell_style()
        for row_index in range(header_row + 1, sheet_rb.nrows):
            room = _merged_or_cell_value(sheet_rb, row_index, room_col)
            if room and room in assignments:
                staff_text = _assignment_text(assignments[room])
                sheet_wb.write(row_index, staff_col, staff_text, staff_style)
                _set_staff_row_height(
                    sheet_wb, row_index, staff_text, staff_col, sheet_rb.merged_cells
                )
    _save_workbook(wb, target_path)
    print(f"[OK] 多场次监考安排已写入: {target_path}")


def write_session_assignments_to_xlsx(schedule_results, input_output_path, header_row=2, output_path=None, add_staff_border=True):
    target_path = output_path or input_output_path
    if output_path and output_path != input_output_path:
        shutil.copyfile(input_output_path, output_path)
    workbook = load_workbook(target_path)
    for sheet in workbook.worksheets:
        result = schedule_results.get(str(sheet.title))
        if not result:
            continue
        headers = [cell.value for cell in sheet[header_row + 1]]
        room_col = _find_column(headers, ["教室编号", "考场号", "room"])
        staff_col = _find_column(headers, ["监考人员", "监考教师", "staff", "proctor"])
        if room_col is None or staff_col is None:
            raise ValueError(f"工作表 {sheet.title} 未找到教室编号或监考人员列")
        previous_room = ""
        for row_index in range(header_row + 2, sheet.max_row + 1):
            room = str(sheet.cell(row_index, room_col + 1).value or "").strip() or previous_room
            if room:
                previous_room = room
            if room in result["assignments"]:
                target = sheet.cell(row_index, staff_col + 1)
                if target.__class__.__name__ == "MergedCell":
                    for merged in sheet.merged_cells.ranges:
                        if target.coordinate in merged:
                            target = sheet.cell(merged.min_row, merged.min_col)
                            break
                _set_xlsx_staff_cell(sheet, target, _assignment_text(result["assignments"][room]), add_staff_border)
    workbook.save(target_path)
    print(f"[OK] 多场次监考安排已写入: {target_path}")


def _merged_or_cell_value(sheet, row_index, col_index):
    for rlo, rhi, clo, chi in sheet.merged_cells:
        if rlo <= row_index < rhi and clo <= col_index < chi:
            return str(sheet.cell_value(rlo, clo)).strip() if row_index == rlo else ""
    value = sheet.cell_value(row_index, col_index)
    return str(value).strip() if value not in (None, "") else ""


def parse_room_groups(rules):
    """解析考场号分组：101-112；201,203,205；C101-C112。"""
    if isinstance(rules, str):
        parts = [part.strip() for part in re.split(r"[;；\n]+", rules) if part.strip()]
    else:
        parts = [str(part).strip() for part in rules if str(part).strip()]
    groups = []
    for part in parts:
        rooms = []
        for token in re.split(r"[,，、\s]+", part):
            if not token:
                continue
            match = re.fullmatch(r"([A-Za-z\u4e00-\u9fff]*)(\d+)\s*-\s*([A-Za-z\u4e00-\u9fff]*)(\d+)", token)
            if match:
                prefix1, start_text, prefix2, end_text = match.groups()
                if prefix1.casefold() != prefix2.casefold():
                    raise ValueError(f"考场号范围前缀不一致: {token}")
                start, end = int(start_text), int(end_text)
                if start > end or end - start > 10000:
                    raise ValueError(f"考场号范围无效: {token}")
                width = max(len(start_text), len(end_text))
                rooms.extend(f"{prefix1}{number:0{width}d}" for number in range(start, end + 1))
            else:
                rooms.append(token)
        if not rooms:
            raise ValueError(f"分组规则为空: {part}")
        groups.append(list(dict.fromkeys(rooms)))
    if not groups:
        raise ValueError("请至少输入一条考场分组规则")
    return groups


def _load_schedule(input_path, header_row):
    rb = xlrd.open_workbook(input_path, formatting_info=True)
    sheet = rb.sheet_by_index(0)
    headers = sheet.row_values(header_row)
    room_col = _find_column(headers, ["教室编号", "考场号", "room"])
    staff_col = _find_column(headers, ["监考人员", "监考教师", "staff", "proctor"])
    if room_col is None or staff_col is None:
        raise ValueError("模板中未找到教室编号/考场号或监考人员列")

    records = []
    previous_room = ""
    previous_staff = ""
    for row_index in range(header_row + 1, sheet.nrows):
        row = _pad_row(sheet.row_values(row_index), len(headers))
        room = _merged_or_cell_value(sheet, row_index, room_col) or previous_room
        staff = _merged_or_cell_value(sheet, row_index, staff_col) or previous_staff
        if room:
            previous_room = room
        if staff:
            previous_staff = staff
        row[room_col] = room
        row[staff_col] = staff
        records.append(row)
    return rb, sheet, headers, room_col, staff_col, records


def get_schedule_rooms(input_path, header_row=2):
    """读取排考结果中的全部教室编号，供分组窗口自动识别范围。"""
    rb = xlrd.open_workbook(input_path, formatting_info=True)
    rooms = []
    for sheet in rb.sheets():
        headers = sheet.row_values(header_row)
        room_col = _find_column(headers, ["教室编号", "考场号", "room"])
        if room_col is None:
            continue
        previous_room = ""
        for row_index in range(header_row + 1, sheet.nrows):
            room = _merged_or_cell_value(sheet, row_index, room_col) or previous_room
            if room:
                previous_room = room
                if room not in rooms:
                    rooms.append(room)
    return rooms


def _export_styles():
    """分组表使用与原模板一致的清晰打印样式。"""
    title = xlwt.easyxf(
        "font: name 微软雅黑, height 220, bold on;"
        "align: horiz center, vert center, wrap on"
    )
    subtitle = xlwt.easyxf(
        "font: name 微软雅黑, height 180;"
        "align: horiz left, vert center, wrap on"
    )
    header = xlwt.easyxf(
        "font: name 微软雅黑, height 180, bold on;"
        "align: horiz center, vert center, wrap on;"
        "borders: left thin, right thin, top thin, bottom thin;"
        "pattern: pattern solid, fore_colour ice_blue"
    )
    body = xlwt.easyxf(
        "font: name 微软雅黑, height 180;"
        "align: horiz center, vert center, wrap on;"
        "borders: left thin, right thin, top thin, bottom thin"
    )
    note = xlwt.easyxf(
        "font: name 微软雅黑, height 180, italic on, colour red;"
        "align: horiz left, vert center"
    )
    return title, subtitle, header, body, note


def _write_rows(ws, rows, header_row, room_col, staff_col, source_sheet=None):
    """写入分组表，同时保留模板的标题、表头、列宽、行高和合并结构。"""
    title, subtitle, header_style, body_style, _ = _export_styles()
    column_count = max(len(row) for row in rows) if rows else 1
    if source_sheet is not None:
        for col_index in range(column_count):
            info = source_sheet.colinfo_map.get(col_index)
            ws.col(col_index).width = info.width if info else 2200
        for row_index in range(len(rows)):
            info = source_sheet.rowinfo_map.get(row_index)
            if info and info.height:
                ws.row(row_index).height = info.height
    _prepare_staff_column(ws, staff_col)

    header_merges = []
    if source_sheet is not None:
        header_merges = [
            (rlo, rhi, clo, chi)
            for rlo, rhi, clo, chi in source_sheet.merged_cells
            if rhi <= header_row + 1
        ]

    for row_index, row in enumerate(rows):
        style = title if row_index == 0 else subtitle if row_index < header_row else header_style
        if row_index > header_row:
            style = body_style
        for col_index in range(column_count):
            if any(rlo <= row_index < rhi and clo <= col_index < chi for rlo, rhi, clo, chi in header_merges):
                continue
            if row_index > header_row and col_index in (room_col, staff_col):
                continue
            value = row[col_index] if col_index < len(row) else ""
            ws.write(row_index, col_index, value, style)

    # 标题及表头的合并单元格直接沿用原模板结构。
    if source_sheet is not None:
        for rlo, rhi, clo, chi in header_merges:
            value = rows[rlo][clo] if rlo < len(rows) and clo < len(rows[rlo]) else ""
            ws.write_merge(rlo, rhi - 1, clo, chi - 1, value, title if rlo == 0 else subtitle if rlo < header_row else header_style)

    data_start = header_row + 1
    for col_index in (room_col, staff_col):
        start = 0
        values = [str(row[col_index]) for row in rows[data_start:]]
        while start < len(values):
            end = start
            while end + 1 < len(values) and values[end + 1] == values[start]:
                end += 1
            first_row = data_start + start
            last_row = data_start + end
            if values[start]:
                if first_row == last_row:
                    ws.write(first_row, col_index, values[start], body_style)
                else:
                    ws.write_merge(first_row, last_row, col_index, col_index, values[start], body_style)
                if col_index == staff_col:
                    _set_staff_row_height(
                        ws, first_row, values[start], staff_col,
                        source_sheet.merged_cells if source_sheet is not None else None
                    )
            start = end + 1


def split_excel_by_room_groups(input_path, output_path, header_row, rules):
    """按照考场号范围/列表生成分组 Sheet，并保留标题行和表头。"""
    groups = parse_room_groups(rules)
    _, source_sheet, headers, room_col, staff_col, records = _load_schedule(input_path, header_row)
    workbook = xlwt.Workbook(encoding="utf-8")
    for index, group in enumerate(groups, start=1):
        group_set = {_room_key(room) for room in group}
        selected = [row for row in records if _room_key(row[room_col]) in group_set]
        rows = [source_sheet.row_values(row_index) for row_index in range(header_row + 1)]
        rows.extend(selected)
        sheet_name = f"第{index}组"[:31]
        worksheet = workbook.add_sheet(sheet_name)
        _write_rows(worksheet, rows, header_row, room_col, staff_col, source_sheet)
        if not selected:
            worksheet.write(header_row + 1, 0, f"未找到匹配考场：{','.join(group)}", _export_styles()[-1])
    _save_workbook(workbook, output_path)
    print(f"[OK] 已按考场号导出 {len(groups)} 组: {output_path}")


def split_schedule_by_room_groups(input_path, output_path, header_row, rules, session_ids=None):
    """多场次模板按考场分组，每个场次和分组生成一个工作表。"""
    if str(input_path).lower().endswith(".xlsx"):
        return _split_schedule_by_room_groups_xlsx(input_path, output_path, header_row, rules, session_ids)
    groups = parse_room_groups(rules)
    rb = xlrd.open_workbook(input_path, formatting_info=True)
    workbook = xlwt.Workbook(encoding="utf-8")
    total = 0
    for sheet in rb.sheets():
        if session_ids and str(sheet.name) not in {str(value) for value in session_ids}:
            continue
        headers = sheet.row_values(header_row)
        room_col = _find_column(headers, ["教室编号", "考场号", "room"])
        staff_col = _find_column(headers, ["监考人员", "监考教师", "staff", "proctor"])
        if room_col is None or staff_col is None:
            continue
        records = []
        previous_room = ""
        previous_staff = ""
        for row_index in range(header_row + 1, sheet.nrows):
            row = _pad_row(sheet.row_values(row_index), len(headers))
            room = _merged_or_cell_value(sheet, row_index, room_col) or previous_room
            staff = _merged_or_cell_value(sheet, row_index, staff_col) or previous_staff
            if room:
                previous_room = room
            if staff:
                previous_staff = staff
            row[room_col] = room
            row[staff_col] = staff
            records.append(row)
        for index, group in enumerate(groups, start=1):
            group_set = {_room_key(room) for room in group}
            selected = [row for row in records if _room_key(row[room_col]) in group_set]
            rows = [sheet.row_values(row_index) for row_index in range(header_row + 1)]
            rows.extend(selected)
            name = f"{sheet.name}-第{index}组"[:31]
            worksheet = workbook.add_sheet(name)
            _write_rows(worksheet, rows, header_row, room_col, staff_col, sheet)
            if not selected:
                worksheet.write(header_row + 1, 0, f"未找到匹配考场：{','.join(group)}", _export_styles()[-1])
            total += 1
    _save_workbook(workbook, output_path)
    print(f"[OK] 多场次分组导出 {total} 个工作表: {output_path}")


def _split_schedule_by_room_groups_xlsx(input_path, output_path, header_row, rules, session_ids=None):
    groups = parse_room_groups(rules)
    source = load_workbook(input_path, data_only=False)
    workbook = Workbook()
    workbook.remove(workbook.active)
    wanted = {str(value) for value in session_ids} if session_ids else None
    total = 0
    for sheet in source.worksheets:
        if wanted and str(sheet.title) not in wanted:
            continue
        headers = [cell.value for cell in sheet[header_row + 1]]
        room_col = _find_column(headers, ["教室编号", "考场号", "room"])
        staff_col = _find_column(headers, ["监考人员", "监考教师", "staff", "proctor"])
        if room_col is None or staff_col is None:
            continue
        rows = []
        previous_room = ""
        previous_staff = ""
        for row_index in range(header_row + 2, sheet.max_row + 1):
            values = _pad_row([sheet.cell(row_index, col).value for col in range(1, sheet.max_column + 1)], len(headers))
            room = str(values[room_col] or "").strip() or previous_room
            staff = str(values[staff_col] or "").strip() or previous_staff
            if room:
                previous_room = room
            if staff:
                previous_staff = staff
            values[room_col] = room
            values[staff_col] = staff
            rows.append((room, values, row_index))
        for index, group in enumerate(groups, start=1):
            output_sheet = workbook.create_sheet(f"{sheet.title}-第{index}组"[:31])
            _copy_xlsx_dimensions(sheet, output_sheet)
            for row_index in range(1, header_row + 2):
                for col_index in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(row_index, col_index)
                    target_cell = output_sheet.cell(row_index, col_index, source_cell.value)
                    _copy_xlsx_cell(source_cell, target_cell)
            for merged in sheet.merged_cells.ranges:
                if merged.max_row <= header_row + 1:
                    output_sheet.merge_cells(str(merged))
            group_set = {_room_key(room) for room in group}
            selected = [(values, source_row) for room, values, source_row in rows if _room_key(room) in group_set]
            for offset, (values, source_row) in enumerate(selected, start=header_row + 2):
                for col_index, value in enumerate(values, start=1):
                    target_cell = output_sheet.cell(offset, col_index, value)
                    _copy_xlsx_cell(sheet.cell(source_row, col_index), target_cell)
            if not selected:
                output_sheet.cell(header_row + 2, 1, f"未找到匹配考场：{','.join(group)}")
            total += 1
    if not total:
        raise ValueError("未找到可分组的考试场次")
    workbook.save(output_path)
    print(f"[OK] 多场次分组导出 {total} 个工作表: {output_path}")


def split_excel_by_serial(input_path, output_path, header_row, sheet_count):
    """兼容旧入口：按原表中的考场顺序平均切组，但不再依赖序号列。"""
    if sheet_count < 1:
        raise ValueError("分组数量必须大于0")
    _, _, _, room_col, _, records = _load_schedule(input_path, header_row)
    rooms = list(dict.fromkeys(row[room_col] for row in records if row[room_col]))
    if not rooms:
        raise ValueError("模板中没有有效考场号")
    size = (len(rooms) + sheet_count - 1) // sheet_count
    groups = [rooms[start:start + size] for start in range(0, len(rooms), size)]
    return split_excel_by_room_groups(input_path, output_path, header_row, groups)
